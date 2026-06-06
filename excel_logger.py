"""
excel_logger.py — Real-time Excel trade logger.

Creates/updates logs/demo_trades.xlsx after EVERY trade.
Includes:
  - Sheet 1: Trade Log (every trade, color-coded wins/losses)
  - Sheet 2: Summary (P&L, win rate, R:R, equity curve)
  - Sheet 3: Signal Details (EMA, RSI, trend info per trade)

Install: pip install openpyxl
"""

import os
import logging
from datetime import datetime
from typing import List
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as xl_numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint

log = logging.getLogger(__name__)
IST = ZoneInfo("Asia/Kolkata")

# ─── Colors ───────────────────────────────────────────────────────────────────
C_NAVY    = "0D1B2A"
C_GREEN   = "00C896"
C_RED     = "FF4D4D"
C_YELLOW  = "FFD700"
C_LGRAY   = "F5F5F5"
C_MGRAY   = "CCCCCC"
C_DGRAY   = "444444"
C_WHITE   = "FFFFFF"
C_ORANGE  = "FF8C00"
C_BLUE    = "1E90FF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_border():
    thick = Side(style="medium")
    thin  = Side(style="thin")
    return Border(left=thick, right=thick, top=thick, bottom=thick)


class ExcelTradeLogger:
    """
    Writes/updates demo_trades.xlsx after every trade close.
    Call update() after each trade, generate_full_report() at session end.
    """

    def __init__(self, path: str = "logs/demo_trades.xlsx", index: str = "NIFTY",
                 capital: float = 500_000):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        self.path    = path
        self.index   = index
        self.capital = capital
        self._ensure_workbook()

    # ─── Public API ───────────────────────────────────────────────────────────

    def log_trade(self, pos) -> None:
        """Append one closed trade to the Excel file immediately."""
        try:
            wb = load_workbook(self.path)
            self._append_trade_row(wb, pos)
            self._refresh_summary(wb, self._get_all_trades(wb))
            wb.save(self.path)
            log.info(f"[XLS] Trade logged → {self.path}")
        except Exception as exc:
            log.error(f"[XLS] log_trade error: {exc}")

    def generate_full_report(self, trades: list) -> str:
        """Rebuild full workbook with all trades + charts."""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            self._build_trade_sheet(wb, trades)
            self._build_summary_sheet(wb, trades)
            self._build_signal_sheet(wb, trades)
            self._build_equity_chart(wb, trades)
            wb.save(self.path)
            log.info(f"[XLS] Full report saved → {self.path}")
            return self.path
        except Exception as exc:
            log.error(f"[XLS] generate_full_report error: {exc}")
            import traceback; traceback.print_exc()
            return ""

    # ─── Sheet Builders ───────────────────────────────────────────────────────

    def _build_trade_sheet(self, wb: Workbook, trades: list):
        ws = wb.create_sheet("Trade Log")

        # ── Title row ────────────────────────────────────────────────────────
        ws.merge_cells("A1:T1")
        ws["A1"] = f"📊 {self.index} Paper Trading Log  |  Generated: {datetime.now(IST).strftime('%d %b %Y %H:%M IST')}"
        ws["A1"].font      = _font(bold=True, color=C_WHITE, size=13)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _align()
        ws.row_dimensions[1].height = 28

        # ── Column headers ────────────────────────────────────────────────────
        headers = [
            ("#",          6),
            ("Date",      12),
            ("Entry Time",12),
            ("Exit Time", 12),
            ("Signal",     9),
            ("Strategy",  18),
            ("Strike",    12),
            ("Expiry",    13),
            ("Lots",       7),
            ("Entry ₹",   11),
            ("Exit ₹",    11),
            ("SL ₹",      10),
            ("TP ₹",      10),
            ("P&L ₹",     14),
            ("P&L %",     10),
            ("Duration",  10),
            ("Exit Reason",14),
            ("EMA9",       9),
            ("EMA15",      9),
            ("Score",      8),
        ]

        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font      = _font(bold=True, color=C_WHITE, size=9)
            cell.fill      = _fill(C_NAVY)
            cell.alignment = _align()
            cell.border    = _header_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

        # ── Data rows ─────────────────────────────────────────────────────────
        cum_pnl = 0
        for i, t in enumerate(trades, 1):
            row = i + 2
            pnl  = t.pnl if hasattr(t, "pnl") else 0
            cum_pnl += pnl
            is_win   = pnl > 0

            # Row fill alternating + win/loss highlight
            if is_win:
                row_fill = _fill("E6F9F2") if i % 2 == 0 else _fill("D0F5E8")
            else:
                row_fill = _fill("FFF0F0") if i % 2 == 0 else _fill("FFE0E0")

            dur = ""
            if hasattr(t, "entry_time") and hasattr(t, "exit_time") and t.exit_time:
                mins = int((t.exit_time - t.entry_time).total_seconds() / 60)
                dur  = f"{mins}m"

            values = [
                i,
                t.entry_time.strftime("%d-%b-%Y") if t.entry_time else "",
                t.entry_time.strftime("%H:%M:%S") if t.entry_time else "",
                t.exit_time.strftime("%H:%M:%S")  if t.exit_time  else "",
                t.cross if hasattr(t,"cross") else getattr(t,"signal_type","?"),
                getattr(t, "mode", ""),
                f"{t.strike}{t.opt_type}" if hasattr(t,"opt_type") else str(getattr(t,"strike","")),
                getattr(t, "expiry", ""),
                getattr(t, "lots", 1),
                round(t.entry_price, 1),
                round(t.exit_price, 1),
                round(t.sl_price, 1),
                round(t.tp_price, 1),
                round(pnl, 2),
                round(t.pnl_pct, 2) if hasattr(t, "pnl_pct") else 0,
                dur,
                getattr(t, "exit_reason", ""),
                round(getattr(t, "e9",  0), 1),
                round(getattr(t, "e15", 0), 1),
                round(getattr(t, "score", 0), 0),
            ]

            for col_idx, val in enumerate(values, 1):
                cell             = ws.cell(row=row, column=col_idx, value=val)
                cell.font        = _font(size=9)
                cell.fill        = row_fill
                cell.alignment   = _align()
                cell.border      = _border()

                # Special formatting
                if col_idx == 14:  # P&L
                    cell.number_format = '₹#,##0.00;[Red](₹#,##0.00)'
                    cell.font = _font(bold=True, color=C_GREEN if is_win else C_RED, size=9)
                if col_idx == 15:  # P&L%
                    cell.number_format = '0.00%;[Red]-0.00%'
                    cell.value = val / 100  # store as decimal for % format
                if col_idx == 5:   # Signal
                    cell.fill = _fill("D0F5E8") if str(val)=="CALL" else _fill("FFE0E0")
                    cell.font = _font(bold=True, color=C_GREEN if str(val)=="CALL" else C_RED, size=9)

        # Add totals row
        if trades:
            tot_row = len(trades) + 3
            ws.merge_cells(f"A{tot_row}:M{tot_row}")
            ws[f"A{tot_row}"]        = "TOTAL"
            ws[f"A{tot_row}"].font   = _font(bold=True, color=C_WHITE, size=10)
            ws[f"A{tot_row}"].fill   = _fill(C_NAVY)
            ws[f"A{tot_row}"].alignment = _align()

            total_pnl = sum(t.pnl for t in trades)
            pnl_cell  = ws.cell(row=tot_row, column=14, value=round(total_pnl, 2))
            pnl_cell.font   = _font(bold=True, color=C_WHITE, size=10)
            pnl_cell.fill   = _fill(C_GREEN if total_pnl >= 0 else C_RED)
            pnl_cell.number_format = '₹#,##0.00'
            pnl_cell.alignment = _align()

        ws.auto_filter.ref = f"A2:T{len(trades)+2}"

    def _build_summary_sheet(self, wb: Workbook, trades: list):
        ws = wb.create_sheet("Summary")

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{self.index} Paper Trading — Session Summary"
        ws["A1"].font      = _font(bold=True, color=C_WHITE, size=14)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _align()
        ws.row_dimensions[1].height = 30

        if not trades:
            ws["A3"] = "No trades taken yet."
            return

        # Compute stats
        pnls   = [t.pnl for t in trades]
        wins   = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p <= 0]
        total  = len(pnls)
        wr     = len(wins)/total*100 if total else 0
        aw     = sum(wins)/len(wins)     if wins   else 0
        al     = sum(losses)/len(losses) if losses else 0
        rr     = abs(aw/al)              if al     else 0
        total_pnl = sum(pnls)
        max_win   = max(pnls)  if pnls else 0
        max_loss  = min(pnls)  if pnls else 0

        # Equity curve: running sum
        cum   = 0
        peak  = 0
        max_dd = 0
        for p in pnls:
            cum  += p
            peak  = max(peak, cum)
            dd    = peak - cum
            max_dd= max(max_dd, dd)

        stats = [
            ("PERFORMANCE METRICS", "", ""),
            ("Total Trades",        total,              ""),
            ("Winning Trades",      len(wins),          ""),
            ("Losing Trades",       len(losses),        ""),
            ("Win Rate",            f"{wr:.1f}%",       ""),
            ("", "", ""),
            ("P&L SUMMARY", "", ""),
            ("Total P&L",           f"₹{total_pnl:+,.0f}", "WIN" if total_pnl>=0 else "LOSS"),
            ("Avg Win",             f"₹{aw:+,.0f}",     ""),
            ("Avg Loss",            f"₹{al:+,.0f}",     ""),
            ("Best Trade",          f"₹{max_win:+,.0f}",""),
            ("Worst Trade",         f"₹{max_loss:+,.0f}",""),
            ("Reward:Risk Ratio",   f"{rr:.2f}",        ""),
            ("Max Drawdown",        f"₹{max_dd:,.0f}",  ""),
            ("", "", ""),
            ("ACCOUNT", "", ""),
            ("Starting Capital",    f"₹{self.capital:,.0f}", ""),
            ("Ending Capital",      f"₹{self.capital+total_pnl:,.0f}", ""),
            ("Total Return",        f"{total_pnl/self.capital*100:+.2f}%", ""),
        ]

        for r, (label, value, tag) in enumerate(stats, 3):
            ws.row_dimensions[r].height = 20
            if label in ("PERFORMANCE METRICS","P&L SUMMARY","ACCOUNT"):
                ws.merge_cells(f"A{r}:F{r}")
                ws[f"A{r}"]            = label
                ws[f"A{r}"].font       = _font(bold=True, color=C_WHITE, size=10)
                ws[f"A{r}"].fill       = _fill(C_DGRAY)
                ws[f"A{r}"].alignment  = _align(h="left")
                continue
            if not label:
                continue

            lc = ws.cell(row=r, column=1, value=label)
            lc.font = _font(bold=True, size=10)
            lc.fill = _fill(C_LGRAY)
            lc.alignment = _align(h="left")
            lc.border = _border()

            vc = ws.cell(row=r, column=2, value=value)
            vc.font = _font(bold=True, size=11,
                            color=C_GREEN if tag=="WIN" else (C_RED if tag=="LOSS" else "000000"))
            vc.alignment = _align()
            vc.border = _border()

            ws.merge_cells(f"B{r}:F{r}")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20

        # Equity curve table
        eq_row = len(stats) + 5
        ws.cell(row=eq_row, column=1, value="EQUITY CURVE").font = _font(bold=True, size=11, color=C_WHITE)
        ws.cell(row=eq_row, column=1).fill = _fill(C_NAVY)
        ws.merge_cells(f"A{eq_row}:F{eq_row}")

        eq_hdrs = ["Trade#","Date","Signal","P&L ₹","Cum P&L ₹","Capital ₹"]
        for c, h in enumerate(eq_hdrs, 1):
            cell = ws.cell(row=eq_row+1, column=c, value=h)
            cell.font = _font(bold=True, color=C_WHITE, size=9)
            cell.fill = _fill(C_NAVY)
            cell.alignment = _align()

        cum = 0
        for i, t in enumerate(trades, 1):
            cum += t.pnl
            r    = eq_row + 1 + i
            row_vals = [
                i,
                t.entry_time.strftime("%d-%b %H:%M") if t.entry_time else "",
                getattr(t,"cross", "?"),
                round(t.pnl, 2),
                round(cum, 2),
                round(self.capital + cum, 2),
            ]
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = _font(size=9, color=C_GREEN if t.pnl>0 else C_RED)
                cell.fill = _fill("E6F9F2" if t.pnl>0 else "FFF0F0")
                cell.alignment = _align()
                cell.border = _border()
                if c in (4,5,6):
                    cell.number_format = '₹#,##0.00'

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18

    def _build_signal_sheet(self, wb: Workbook, trades: list):
        ws = wb.create_sheet("Signal Details")

        ws.merge_cells("A1:L1")
        ws["A1"] = "Entry Signal Details — Per Trade Indicator Values"
        ws["A1"].font      = _font(bold=True, color=C_WHITE, size=12)
        ws["A1"].fill      = _fill(C_NAVY)
        ws["A1"].alignment = _align()
        ws.row_dimensions[1].height = 26

        hdrs = ["#","Time","Signal","Strike","Mode","EMA9(3m)","EMA15(3m)",
                "RSI(3m)","Trend 15m","Trend 1h","VIX","Entry Reasons"]
        widths = [6,12,10,12,22,12,12,10,12,10,8,60]

        for c, (h, w) in enumerate(zip(hdrs, widths), 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font      = _font(bold=True, color=C_WHITE, size=9)
            cell.fill      = _fill(C_NAVY)
            cell.alignment = _align()
            cell.border    = _header_border()
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.freeze_panes = "A3"

        for i, t in enumerate(trades, 1):
            row  = i + 2
            bg   = _fill("E6F9F2") if t.pnl>0 else _fill("FFF0F0")
            vals = [
                i,
                t.entry_time.strftime("%H:%M:%S") if t.entry_time else "",
                getattr(t,"cross","?"),
                f"{t.strike}{t.opt_type}" if hasattr(t,"opt_type") else str(t.strike),
                getattr(t,"mode",""),
                round(getattr(t,"e9",  0), 1),
                round(getattr(t,"e15", 0), 1),
                round(getattr(t,"rsi", 0), 1),
                getattr(t,"trend_15m",""),
                getattr(t,"trend_1h", ""),
                round(getattr(t,"vix", 0), 2),
                " | ".join(getattr(t,"reasons",t.entry_reasons if hasattr(t,"entry_reasons") else [])[:3]),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font      = _font(size=9)
                cell.fill      = bg
                cell.alignment = _align(h="left" if c==12 else "center", wrap=c==12)
                cell.border    = _border()
            ws.row_dimensions[row].height = 18

    def _build_equity_chart(self, wb: Workbook, trades: list):
        if not trades:
            return
        ws = wb.create_sheet("Equity Chart")

        ws["A1"] = "Trade #"
        ws["B1"] = "Cumulative P&L (₹)"
        ws["C1"] = "Capital (₹)"
        ws["A1"].font = ws["B1"].font = ws["C1"].font = _font(bold=True, color=C_WHITE, size=10)
        ws["A1"].fill = ws["B1"].fill = ws["C1"].fill = _fill(C_NAVY)

        cum = 0
        for i, t in enumerate(trades, 1):
            cum += t.pnl
            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=round(cum, 2))
            ws.cell(row=i+1, column=3, value=round(self.capital + cum, 2))

        # Line chart
        chart = LineChart()
        chart.title  = f"{self.index} Paper Trading — Equity Curve"
        chart.style  = 10
        chart.y_axis.title = "Cumulative P&L (₹)"
        chart.x_axis.title = "Trade Number"
        chart.width  = 25
        chart.height = 14

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(trades)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = C_GREEN
        chart.series[0].graphicalProperties.line.width     = 20000

        ws.add_chart(chart, "E2")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22

    # ─── Helpers ──────────────────────────────────────────────────────────────

    def _ensure_workbook(self):
        """Create blank workbook if file doesn't exist."""
        if not os.path.exists(self.path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Trade Log"
            ws["A1"] = "Waiting for first trade..."
            ws["A1"].font = _font(bold=True, size=12)
            wb.save(self.path)
            log.info(f"[XLS] Created: {self.path}")

    def _append_trade_row(self, wb: Workbook, pos):
        """Quick append to Trade Log sheet (for real-time logging)."""
        if "Trade Log" not in wb.sheetnames:
            wb.create_sheet("Trade Log")
        ws  = wb["Trade Log"]
        row = ws.max_row + 1
        pnl = pos.pnl if hasattr(pos,"pnl") else 0

        vals = [
            row - 1,
            pos.entry_time.strftime("%d-%b-%Y") if pos.entry_time else "",
            pos.entry_time.strftime("%H:%M:%S") if pos.entry_time else "",
            pos.exit_time.strftime("%H:%M:%S")  if pos.exit_time  else "",
            getattr(pos,"cross", getattr(pos,"signal_type",{}).value if hasattr(getattr(pos,"signal_type",None),"value") else "?"),
            getattr(pos,"mode",""),
            f"{pos.strike}{pos.opt_type}" if hasattr(pos,"opt_type") else str(pos.strike),
            getattr(pos,"expiry",""),
            getattr(pos,"lots",1),
            round(pos.entry_price, 1),
            round(pos.exit_price,  1),
            round(pos.sl_price,    1),
            round(pos.tp_price,    1),
            round(pnl, 2),
            round(pos.pnl_pct, 2) if hasattr(pos,"pnl_pct") else 0,
            getattr(pos,"exit_reason",""),
            round(getattr(pos,"e9",  0), 1),
            round(getattr(pos,"e15", 0), 1),
            round(getattr(pos,"score",0), 0),
        ]

        is_win  = pnl > 0
        row_fill= _fill("D0F5E8") if is_win else _fill("FFE0E0")

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _font(size=9,
                              bold=(c==14),
                              color=C_GREEN if (c==14 and is_win) else (C_RED if (c==14 and not is_win) else "000000"))
            cell.fill      = row_fill
            cell.alignment = _align()
            cell.border    = _border()

    def _get_all_trades(self, wb: Workbook) -> list:
        """Not needed for quick append — summary uses live trade list."""
        return []

    def _refresh_summary(self, wb: Workbook, trades: list):
        """Placeholder — full summary is built at session end."""
        pass